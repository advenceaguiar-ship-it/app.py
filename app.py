import streamlit as st
from openai import OpenAI
import json
import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from gtts import gTTS
import io

st.set_page_config(page_title="Assistente de Manutenção", layout="wide")

st.title("🛠️ Assistente de Manutenção Industrial (Zero-UI)")
st.write("Fale com a IA. O relatório, tempo de serviço e checklist são validados automaticamente.")

# --- CONTROLE DE ESTADO BLINDADO ---
if "etapa" not in st.session_state:
    st.session_state.etapa = 1
if "dados_parciais" not in st.session_state:
    st.session_state.dados_parciais = {}
if "texto_ia" not in st.session_state:
    st.session_state.texto_ia = ""
if "audio_bytes_ia" not in st.session_state:
    st.session_state.audio_bytes_ia = None

def reiniciar_os():
    st.session_state.etapa = 1
    st.session_state.dados_parciais = {}
    st.session_state.texto_ia = ""
    st.session_state.audio_bytes_ia = None
    st.rerun()

# --- CAMINHO DO EXCEL SEGURO ---
arquivo_excel = "relatorios_manutencao.xlsx"

def salvar_excel_seguro(df_nova, caminho):
    if not os.path.exists(caminho):
        df_nova.to_excel(caminho, index=False)
    else:
        try:
            df_existente = pd.read_excel(caminho)
            df_final = pd.concat([df_existente, df_nova], ignore_index=True)
            df_final.to_excel(caminho, index=False)
        except PermissionError:
            caminho_alt = caminho.replace(".xlsx", "_copia_segura.xlsx")
            df_nova.to_excel(caminho_alt, index=False)
            st.warning(f"⚠️ O Excel principal está aberto. Salvo temporariamente em: {caminho_alt}")
            return

    try:
        wb = openpyxl.load_workbook(caminho)
        ws = wb.active
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border
                if cell.column in [1, 2, 3, 4]: 
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        ws.row_dimensions[1].height = 28
        wb.save(caminho)
    except Exception:
        pass

# --- BARRA LATERAL (CONFIGURAÇÕES E FILTROS) ---
st.sidebar.header("⚙️ Configurações & Filtros")

if "GROQ_API_KEY" in st.secrets:
    raw_key = st.secrets["GROQ_API_KEY"]
else:
    raw_key = st.sidebar.text_input("Cole a sua API Key da Groq:", type="password")

st.sidebar.divider()
st.sidebar.subheader("🔍 Filtrar Ordens de Serviço")
filtro_tipo = st.sidebar.selectbox("Filtrar por Tipo de Serviço:", ["Todos", "Elétrico", "Mecânico", "Resina"])
filtro_setor = st.sidebar.text_input("Filtrar por Setor (Ex: Cobre, Ferro, Usinagem):")

if raw_key:
    api_key = raw_key.strip()
    client = OpenAI(api_key=api_key, base_url="https://api.groq.com/openai/v1")
    
    if st.sidebar.button("🔄 Cancelar / Iniciar Nova O.S."):
        reiniciar_os()

    # ==========================================
    # ETAPA 1: RELATO INICIAL
    # ==========================================
    if st.session_state.etapa == 1:
        st.subheader("Fase 1: Descreva o problema, serviço e tempo gasto")
        
        audio_file = st.audio_input("Clique no microfone para gravar:")

        if audio_file is not None:
            st.info("🎙️ Áudio capturado com sucesso! Clique no botão abaixo para processar e gerar o relatório.")
            
            if st.button("🚀 Processar Relatório Falado", type="primary"):
                with st.spinner("Transcrevendo e analisando dados com IA..."):
                    transcript = client.audio.transcriptions.create(model="whisper-large-v3", file=audio_file)
                    
                    prompt = f"""Analise o relato do técnico e extraia em formato JSON exatamente com as chaves:
                    - tipo_servico (Estritamente: 'Elétrico', 'Mecânico' ou 'Resina'. Se não se enquadrar, coloque 'Não informada')
                    - setor (Inicie com letra maiúscula)
                    - equipamento (Inicie com letra maiúscula)
                    - falha (Escreva formalmente, iniciando com letra maiúscula)
                    - causa (Se o técnico mencionou a causa, preencha formatando corretamente. Senão, coloque 'Não informada')
                    - acao (Escreva formalmente, iniciando com letra maiúscula)
                    - tempo_gasto (Ex: '7h às 8h', '40 minutos', '2 horas'. Se não citar, coloque 'Não informado')
                    - checklist_seguranca (Analise se o técnico informou no áudio se deixou o setor limpo, seguro, deu baixa, guardou ferramentas e testou. Se disse que fez tudo, preencha 'Concluído no relato'. Senão, 'Pendente')
                    
                    Texto do técnico: {transcript.text}"""
                    
                    response = client.chat.completions.create(
                        model="llama-3.3-70b-versatile",
                        messages=[{"role": "user", "content": prompt}],
                        response_format={"type": "json_object"}
                    )
                    
                    dados = json.loads(response.choices[0].message.content)
                    st.session_state.dados_parciais = dados
                    
                    causa_ok = dados.get("causa") != 'Não informada'
                    checklist_ok = dados.get("checklist_seguranca") == 'Concluído no relato'
                    
                    if causa_ok and checklist_ok:
                        agora = datetime.now().strftime("%d/%m/%Y %H:%M")
                        num_os = 1
                        if os.path.exists(arquivo_excel):
                            try:
                                df_existente = pd.read_excel(arquivo_excel)
                                num_os = len(df_existente) + 1
                            except:
                                pass
                        
                        nova_os = {
                            "Nº O.S.": f"OS-{num_os:03d}",
                            "Data/Hora": agora,
                            "Tipo de Serviço": dados.get("tipo_servico", "Não informada"),
                            "Setor / Área": dados.get("setor", "Não informada"),
                            "Equipamento": dados.get("equipamento", "Não informada"),
                            "Falha Relatada": dados.get("falha", "Não informada"),
                            "Causa Raiz": dados.get("causa", "Não informada"),
                            "Ação Tomada": dados.get("acao", "Não informada"),
                            "Tempo Gasto": dados.get("tempo_gasto", "Não informado"),
                            "Checklist & Limpeza": "Tudo OK (Informado no relato inicial)"
                        }

                        df_nova = pd.DataFrame([nova_os])
                        salvar_excel_seguro(df_nova, arquivo_excel)
                        
                        st.success("✅ Relatório 100% completo! Salvo no Excel com sucesso.")
                        st.balloons()
                        import time
                        time.sleep(2)
                        reiniciar_os()
                    else:
                        texto_fala = "Relatório parcial capturado. "
                        if not causa_ok:
                            texto_fala += "Por favor, informe a causa raiz do problema. "
                        texto_fala += "Confirme também o checklist: o setor foi limpo, o serviço está seguro, deu baixa no almoxarifado, guardou ferramentas e testou? Grave sua resposta."

                        audio_bytes = None
                        try:
                            tts = gTTS(text=texto_fala, lang='pt')
                            fp = io.BytesIO()
                            tts.write_to_fp(fp)
                            fp.seek(0)
                            audio_bytes = fp.read()
                        except Exception:
                            pass

                        st.session_state.texto_ia = texto_fala
                        st.session_state.audio_bytes_ia = audio_bytes
                        st.session_state.etapa = 2
                        st.rerun()

    # ==========================================
    # ETAPA 2: COMPLEMENTO POR VOZ
    # ==========================================
    elif st.session_state.etapa == 2:
        st.subheader("Fase 2: Complemento de Causa e Checklist de Segurança")
        st.warning(f"🤖 **A IA está solicitando:** {st.session_state.texto_ia}")
        
        if st.session_state.audio_bytes_ia is not None:
            st.audio(st.session_state.audio_bytes_ia, format="audio/mp3", autoplay=True)
        
        audio_resposta = st.audio_input("Grave sua resposta complementar:")

        if audio_resposta is not None:
            st.info("🎙️ Resposta gravada! Clique no botão abaixo para concluir a O.S.")
            
            if st.button("🚀 Enviar Resposta e Finalizar O.S.", type="primary"):
                with st.spinner("Processando complemento e salvando no Excel..."):
                    transcript2 = client.audio.transcriptions.create(model="whisper-large-v3", file=audio_resposta)
                    
                    prompt_final = f"""
                    JSON original: {json.dumps(st.session_state.dados_parciais)}
                    Resposta complementar do técnico: {transcript2.text}
                    
                    Tarefa 1: Preencha o campo 'causa' formatando com primeira letra maiúscula se estiver 'Não informada'.
                    Tarefa 2: Crie a chave 'checklist_seguranca' consolidando a situação (Limpeza, segurança, baixa no almoxarifado, ferramentas e teste) formatado e correto.
                    Retorne apenas o JSON final atualizado.
                    """
                    
                    response_final = client.chat.completions.create(
                        model="llama-3.3-70b-versatile",
                        messages=[{"role": "user", "content": prompt_final}],
                        response_format={"type": "json_object"}
                    )
                    
                    dados_finais = json.loads(response_final.choices[0].message.content)
                    
                    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
                    num_os = 1
                    if os.path.exists(arquivo_excel):
                        try:
                            df_existente = pd.read_excel(arquivo_excel)
                            num_os = len(df_existente) + 1
                        except:
                            pass
                    
                    nova_os = {
                        "Nº O.S.": f"OS-{num_os:03d}",
                        "Data/Hora": agora,
                        "Tipo de Serviço": dados_finais.get("tipo_servico", "Não informada"),
                        "Setor / Área": dados_finais.get("setor", "Não informada"),
                        "Equipamento": dados_finais.get("equipamento", "Não informada"),
                        "Falha Relatada": dados_finais.get("falha", "Não informada"),
                        "Causa Raiz": dados_finais.get("causa", "Não informada"),
                        "Ação Tomada": dados_finais.get("acao", "Não informada"),
                        "Tempo Gasto": dados_finais.get("tempo_gasto", "Não informado"),
                        "Checklist & Limpeza": dados_finais.get("checklist_seguranca", "Concluído via voz")
                    }

                    df_nova = pd.DataFrame([nova_os])
                    salvar_excel_seguro(df_nova, arquivo_excel)
                    
                    st.success("✅ Ordem de serviço finalizada e salva com sucesso no Excel!")
                    st.balloons()
                    
                    import time
                    time.sleep(2)
                    reiniciar_os()

    # --- HISTÓRICO COM FILTROS DE PESQUISA ---
    st.divider()
    st.subheader("📋 Histórico de Ordens de Serviço (Filtrado)")
    
    if os.path.exists(arquivo_excel):
        try:
            df_historico = pd.read_excel(arquivo_excel)
            df_filtrado = df_historico.copy()
            
            if filtro_tipo != "Todos":
                df_filtrado = df_filtrado[df_filtrado["Tipo de Serviço"].astype(str).str.contains(filtro_tipo, case=False, na=False)]
                
            if filtro_setor.strip():
                termo_busca = filtro_setor.strip()
                df_filtrado = df_filtrado[df_filtrado["Setor / Área"].astype(str).str.contains(termo_busca, case=False, na=False)]
                
            st.dataframe(df_filtrado, use_container_width=True)
            
            with open(arquivo_excel, "rb") as f:
                st.download_button(
                    label="📥 Baixar Planilha Excel Atualizada",
                    data=f,
                    file_name="relatorios_manutencao.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception:
            st.write("A planilha está aberta no Excel. Feche-a temporariamente se quiser visualizar o histórico na tela.")
    else:
        st.write("Nenhuma O.S. gravada ainda.")

else:
    st.warning("⚠️ Insira a sua chave de API da Groq na barra lateral ou nos Secrets para começar.")
